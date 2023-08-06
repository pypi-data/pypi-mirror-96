import re
import xlrd
import openpyxl
import sys
import os
import shutil
import psutil
import unicodedata
from docx import Document
from openpyxl.utils import column_index_from_string
from datetime import datetime

def EmailFormatValidator(email):
    """
     Validates if an email address is in the correct format and returns a
       PASS (email in correct format) or FAIL (email not in correct format)
    """
    if re.match("[\.\w]{2,}[@]\w+[.]\w+",email,re.IGNORECASE):
    	 return "PASS"
    else:
         return "FAIL"

def ContainsOnlyDigits(str_value):
    """
     Validates if a value contains only digits
       Returns PASS (str_value contains only digits) or FAIL (str_value does not
       contain only digits)
    """
    # Using regex() 
    if re.match('^[0-9]*$', str_value): 
       return   'PASS' 
    else: 
       return   'FAIL'

def ConvertStringToList(string): 
    """
     Converts a string of values to an array to be used in a list variable
    """
    ConvLst = list(string.split(" ")) 
    return ConvLst 

def GetValueInString(value, str): 
    """
     Returns the string value (value) passed in if it is contained within the string (str) or Returns FAIL if the string value (value)
       is not contained within the string (str).
    """
    value = value.lower()
    str = str.lower()
    Match = re.findall(value, str)
    if Match: 
       return    Match
    else: 
       return    'FAIL'
       
def StringFormatValidator(field_format, field_value): 
    """
     Returns PASS if the field_value matches a specified field format (field_format)
     Returns FAIL if the field_value does not match a specified field Regex format (field_format)
   
     Use Example:    ${PASS_FAIL}=    String Format Validator   ^[0-9]{6}-[0-9]{1}$     ${Str}
                 Note: must be any 6 digits (from 0 to 9) dash any one digit (from 0 to 9)  :
                       016349-0 ; 999999-9 ; 000000-0
                        
    """
    Regex_Match = re.match(field_format, field_value)
    if Regex_Match: 
       return    'PASS'
    else: 
       return    'FAIL'

def GetStringPosition(str,fnd): 
    """
    Returns the index of the position of a the first occurrence of a string (fnd) contained in the string value passed in (str)
    """
    str = str.lower()
    fnd = fnd.lower()
    pos = 0
    ind = str.find(fnd)
    pos += str.find(fnd)
    return pos

def GetUniqueItems(list_values):
  """
  Returns the unique list of values from parameter list_values
  """
  for x in list_values:
      if list_values.count(x) > 1:
         list_values.remove(x)
  return list_values

def CountOccurenceOfValueInString(string,sub_string):
    """
     Returns the count of the number of times a value appears in a string
    """
    string = string.lower()
    sub_string = sub_string.lower()
    l=len(sub_string)
    count=0
    for i in range(len(string)-len(sub_string)+1):
        if(string[i:i+len(sub_string)] == sub_string ):      
            count+=1
    return count     

def  KillProcess(process_name):
     """
      Kill the inputted process_name
     """
     for proc in psutil.process_iter():
        # check whether the process_name matches a running process
        if proc.name() == process_name:
            proc.kill()
            
def RemoveSpecialCharacters(str):
     """
     Removes special characters from a string
     """
     alphanumeric = ""
     for character in str:
        if character.isalnum():
           alphanumeric += character
     return	alphanumeric 


def CreateNewWorkbook(filename):
    """
    Creates a workbook object 
    """
    wb = openpyxl.Workbook()
    #Create Sheet and sheetname
    ws = wb.create_sheet("Sheet Name", 0)
    wb.save(filename)    
    
    
def OpenWorkbook(filename):
    """
    Opens a workbook
    """
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
        print(sheet.title)

def GetTestCaseDataRow(testnme, excel_col, fileName, sheetname) :
    """
    Returns a row of data from an excel file for a given test case name
    """
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1
    colEndIndex = worksheet.ncols - 1
    rowStartIndex = 1
    colStartIndex = 0
    testData = []
    dataRow = [] 
    cur_row = 0
    excel_col = int(excel_col)
    Found = 0
    for i in range(rowEndIndex):  
        cur_row = cur_row+1  
        testvalue = worksheet.cell_value(cur_row, excel_col)        
        if testvalue == testnme: 
            Found = 1
            break       
    if (Found == 0):
        return None  	    	                 
    cur_col = colStartIndex
    while cur_col <= colEndIndex:
         cell_type = worksheet.cell_type(cur_row, cur_col)
         value = worksheet.cell_value(cur_row, cur_col)
         dataRow.append(value)
         cur_col+=1              	     
    return dataRow
           

def GetDataRowCount(filename, sheetname) :
    """
    Returns the number rows in a particular sheetname of an excel file
    """
    workbook = xlrd.open_workbook(filename)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1   
    return rowEndIndex

def GetDataByRowIndex(excel_row, filename, sheetname) :
    """
    Return a row of data by the excel_row index passed in
    """
    workbook = openpyxl.load_workbook(filename)
    worksheet =  workbook.get_sheet_by_name(sheetname)
    data_row = []                
    excel_row = int(excel_row)   
    for row in worksheet.iter_rows(min_row=excel_row, max_row=excel_row):
        for cell in row:
           #Append column values to the data row list
           data_row.append(cell.value)                                           	     
    return data_row  # return the row of test data
   
def GetNextAvailableDataRow(filename, sheetname, used_col_letter):
     """
     Returns that next available row of data that is not marked as 'Used' in the column letter (example: A,B,C,etc.) 
        that is passed in 
     """
     wb = openpyxl.load_workbook(filename)
     ws =  wb.get_sheet_by_name(sheetname)
     data_row = []                
     excel_col_number = column_index_from_string(used_col_letter)  
     i = 1
     for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
         i = i + 1 
         if ws.cell(i, excel_col_number).value != "Used": 
           available_row = i
           break # exit for loop
     for row in ws.iter_rows(min_row=available_row, max_row=available_row):               
         for cell in row:
           #Append column values to the data row list
           data_row.append(cell.value)  
     ws.cell(row=available_row, column=excel_col_number).value = "Used" # Update 'Available Row' column cell value to 'Used'
     wb.save(filename) # Save the workbook                                                	     
     return data_row  # return the row of test data

def GetAllDataFromExcelSheet(fileName, sheetname) :
    """
    Returns all of the rows of data from a particular sheetname of an excel file
    """
    workbook = xlrd.open_workbook(fileName)
    worksheet = workbook.sheet_by_name(sheetname)
    rowEndIndex = worksheet.nrows - 1
    colEndIndex = worksheet.ncols - 1
    rowStartIndex = 1
    colStartIndex = 0
    testData = []
    dataRows = []
    curr_row = rowStartIndex
    while curr_row <= rowEndIndex:
         cur_col = colStartIndex
         while cur_col <= colEndIndex:
             cell_type = worksheet.cell_type(curr_row, cur_col)

             value = worksheet.cell_value(curr_row, cur_col)
             dataRows.append(value)
             cur_col+=1
         curr_row += 1
         testData.append(dataRows)
    return dataRows

def WriteToExcelFile(filename, sheetname, data_value, row_index, col_index):

     """
     Write a value into a cell in the worksheet row and column
        that is passed in to the function
     """

     wb = openpyxl.load_workbook(filename)
     ws =  wb.get_sheet_by_name(sheetname)

     r = int(row_index)
     c = int(col_index)


     ws.cell(row=r, column=c).value = ''
     ws.cell(row=r, column=c).value = data_value    #enter the data_value in cell row=r and column=c

     wb.save(filename) # Save the workbook

def ReplaceAccents(text):

    """
    Replaces French accents with the same characters but without accents 
    """

    try:
        text = unicode(text, 'utf-8')
    except NameError: # unicode is a default on python 3
        pass
    text = unicodedata.normalize('NFD', text)\
           .encode('ascii', 'ignore')\
           .decode("utf-8")
    return str(text)


def ReadWordFile(path): 

    """
    Reads the text from a MS Word file and returns the text 
    """  
    document = Document(path)
    for para in document.paragraphs:
       return (para.text)