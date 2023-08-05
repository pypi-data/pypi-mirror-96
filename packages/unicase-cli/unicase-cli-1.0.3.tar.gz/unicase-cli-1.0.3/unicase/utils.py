#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author : mocobk
# @Email : mailmzb@qq.com
# @Time : 2021/2/19 18:20
import json
import pickle
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urljoin

import click
import requests
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.reader import excel
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

excel.warnings.simplefilter('ignore')

ROOT_DIR = Path(__file__).parent
TEMPLATE_XLS = ROOT_DIR.joinpath('template', 'case.xlsm')
TAPD_BASE_URL = 'http://9.134.187.9:8080'

BVT_TASK_TEMPLATE = ROOT_DIR.joinpath('template', 'bvt_task.html')


def secure_filename(filename: str):
    chars_map = {'\\': '_',
                 '/': '_',
                 ':': '_',
                 '*': '_',
                 '?': '',
                 '"': '',
                 '<': '_',
                 '>': '_',
                 '|': '_'}

    trans_table = filename.maketrans(chars_map)
    return filename.translate(trans_table)


class Config:
    CONFIG_FILE = ROOT_DIR.joinpath('config.pkl')

    def get(self):
        if self.CONFIG_FILE.exists():
            return pickle.load(self.CONFIG_FILE.open('rb'))
        else:
            return {}

    def put(self, config):
        pickle.dump(config, self.CONFIG_FILE.open('wb'))


class Tapd:
    def __init__(self, iteration_id=None, base_url=None):
        self.base_url = base_url or TAPD_BASE_URL
        self.iteration_id = iteration_id

    def set_iteration_id(self, iteration_id):
        self.iteration_id = iteration_id

    def get_tapd_iterations(self):
        url = urljoin(self.base_url, '/tapd/origin/iterations')
        response = requests.get(url, params={'status': 'open', 'fields': 'id,name'})
        data = [(item['Iteration']['name'], item['Iteration']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data[:8])

    def get_tapd_iteration_developers(self):
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        response = requests.get(
            url, params={'iteration_id': self.iteration_id, 'fields': 'owner', 'order': 'owner asc'}
        )
        data = [item['Task']['owner'] for item in response.json().get('data', []) if item['Task']['owner']]
        return set(data)

    def get_tapd_iteration_stories(self):
        url = urljoin(self.base_url, '/tapd/origin/stories')
        response = requests.get(url, params={'iteration_id': self.iteration_id, 'fields': 'id,name'})
        data = [(item['Story']['name'], item['Story']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data)

    def get_tapd_case_categories(self):
        url = urljoin(self.base_url, '/tapd/origin/tcase_categories')
        response = requests.get(url, params={'parent_id': 0, 'fields': 'id,name', 'order': 'name asc'})
        data = [(item['TcaseCategory']['name'], item['TcaseCategory']['id']) for item in
                response.json().get('data', [])]
        return OrderedDict(data)

    def create_tapd_task(self, **kwargs):
        """name, description, creator, owner, story_id, iteration_id"""
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        kwargs.setdefault('iteration_id', self.iteration_id)
        kwargs.setdefault('priority', 4)
        kwargs.setdefault('custom_field_one', '冒烟')
        response = requests.post(url, data=kwargs)
        task_id = response.json().get('data', {}).get('Task', {}).get('id')
        return task_id


class Excel:
    def __init__(self, filename, tapd: Tapd = None):
        self.filename = filename
        self.wb = load_workbook(filename, keep_vba=True)
        self.case_ws = self.wb.worksheets[0]
        self.validate_ws = None
        self.case_title_column_letters = self.get_title_column_letter_map(self.case_ws, 2)
        self.tapd = tapd

    @staticmethod
    def get_title_column_letter_map(sheet: Worksheet, title_row=1):
        column_letter_map = {}
        for (cell,) in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=title_row, max_row=title_row):
            if cell.value is None:
                continue
            column_letter_map[cell.value] = cell.column_letter
        return column_letter_map

    def update_case_title_column_letters(self):
        self.case_title_column_letters = self.get_title_column_letter_map(self.case_ws, 2)

    def set_sheet_name(self, name):
        self.case_ws.title = name

    def set_iteration_meta_info(self, data: dict):
        """在用例表中存储额外的元信息, 本想放在sheet表标题的，但标题有长度限制 31"""
        self.case_ws['A1'] = json.dumps(data, ensure_ascii=False)

    def get_iteration_meta_info(self):
        """在用例表中存储额外的元信息, 本想放在sheet表标题的，但标题有长度限制 31"""
        meta_info = self.case_ws['A1'].value
        try:
            return json.loads(meta_info)
        except (json.JSONDecodeError, TypeError):
            return None

    @staticmethod
    def append_cols(sheet, column_data):
        column_index = cur_max_column = sheet.max_column
        if sheet.cell(1, cur_max_column).value is not None:
            column_index += 1
        for index, value in enumerate(column_data, start=1):
            sheet.cell(index, column_index, value)

    def set_data_validation(self, cell, formula_data, show_error_message=True):
        """
        :param cell: 'A1:A10'
        :param formula_data: ['高', '中', '低']
        :param show_error_message: 填错是否弹窗提示
        :return:
        """
        self.append_cols(self.validate_ws, formula_data)
        column_letter = self.validate_ws.cell(1, self.validate_ws.max_column).column_letter
        dv = DataValidation(
            type="list",
            formula1=f"{quote_sheetname(self.validate_ws.title)}!${column_letter}$2:${column_letter}${len(formula_data)}",
            allow_blank=True,
            showErrorMessage=show_error_message)
        dv.error = '请使用下拉选择可用的值'
        dv.errorTitle = '输入的值有误'
        dv.add(cell)
        self.case_ws.add_data_validation(dv)

    def _set_categories_data_validation(self):
        """设置 一级模块（用例目录）的数据验证"""
        column_letter = self.case_title_column_letters['一级模块']
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('一级模块', *self.tapd.get_tapd_case_categories().keys()))

    def _set_developer_data_validation(self):
        """设置 开发人员 的数据验证"""
        column_letter = self.case_title_column_letters['开发人员']
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('开发人员', *self.tapd.get_tapd_iteration_developers()), show_error_message=False)

    def _set_stories_data_validation(self):
        """设置 需求 的数据验证"""
        column_letter = self.case_title_column_letters['需求']
        cell = f'{column_letter}3:{column_letter}4'
        self.set_data_validation(cell, ('需求', *self.tapd.get_tapd_iteration_stories().keys()))
        self.append_cols(self.validate_ws, ('需求ID', *self.tapd.get_tapd_iteration_stories().values()))

    def set_iteration_data_validation(self):
        data_validation_sheet_name = '数据验证（勿删）'
        if data_validation_sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[data_validation_sheet_name])

        self.validate_ws = self.wb.create_sheet('数据验证（勿删）')
        self._set_categories_data_validation()
        self._set_developer_data_validation()
        self._set_stories_data_validation()

    def save(self, filename):
        self.wb.save(filename)


class Cases:
    def __init__(self, excel: Excel):
        self.excel = excel
        self.case_ws = self.excel.case_ws
        self.iteration_meta_info = self.excel.get_iteration_meta_info()
        self.title_column_letters = self.excel.case_title_column_letters

    def validate_data(self):
        pass

    def upload(self, creator):
        pass


class BVTCases(Cases):
    def validate_data(self):
        results = []
        cases_count = 0
        for row_index in range(3, self.case_ws.max_row + 1):
            case_title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
            if self.case_ws[case_title_cell].value:
                priority_cell = f'{self.title_column_letters["用例等级"]}{row_index}'
                if self.case_ws[priority_cell].value == '高':  # 冒烟用例
                    story_cell = f'{self.title_column_letters["需求"]}{row_index}'
                    developer_cell = f'{self.title_column_letters["开发人员"]}{row_index}'
                    if not all([self.case_ws[developer_cell].value, self.case_ws[story_cell].value]):
                        results.append(f'第 {row_index} 行冒烟用例【开发人员】和【需求】不能为空')
                    else:
                        cases_count += 1
        return results, cases_count

    def upload(self, creator):
        # 插入 task_id 列至最后
        task_id_cell = self.case_ws.cell(2, self.case_ws.max_column + 1, 'task_id')

        template = Template(BVT_TASK_TEMPLATE.read_text())
        stories_map = self.iteration_meta_info.get('iteration_stories', {})

        # self.excel.tapd.create_tapd_task = lambda *args, **kwargs: time.sleep(1)

        # 使用进度条
        with click.progressbar(
                iterable=range(3, self.case_ws.max_row + 1),  # 遍历行号
                label="上传进度",
                bar_template="%(label)s | %(bar)s | %(info)s",
                fill_char=click.style("█", fg="green"),
                empty_char=" ",
        ) as bar:
            for row_index in bar:
                developer_cell = f'{self.title_column_letters["开发人员"]}{row_index}'
                if self.case_ws[developer_cell].value:
                    title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
                    precondition_cell = f'{self.title_column_letters["前置条件"]}{row_index}'
                    steps_cell = f'{self.title_column_letters["用例步骤"]}{row_index}'
                    expected_cell = f'{self.title_column_letters["预期结果"]}{row_index}'
                    story_cell = f'{self.title_column_letters["需求"]}{row_index}'
                    task_id = self.excel.tapd.create_tapd_task(
                        name=f'【冒烟用例】{self.case_ws[title_cell].value}',
                        creator=creator,
                        owner=self.case_ws[developer_cell].value,
                        story_id=stories_map.get(self.case_ws[story_cell].value),
                        iteration_id=self.iteration_meta_info['iteration_id'],
                        description=template.render(
                            precondition=self.case_ws[precondition_cell].value,
                            case_steps=self.case_ws[steps_cell].value,
                            expected=self.case_ws[expected_cell].value
                        )
                    )
                    if task_id:
                        self.case_ws.cell(row_index, task_id_cell.column, task_id)
