# Copyright 2020 Karlsruhe Institute of Technology
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import os
import sys

from kadi_apy import apy_command
from kadi_apy.cli.commons import validate_metadatum
from openpyxl import load_workbook
from xmlhelpy import argument
from xmlhelpy import Choice
from xmlhelpy import option
from xmlhelpy import Path
from xmlhelpy import String

from .main import converter


def _read_value(ws, letter, number):
    value = ws[f"{letter}{number}"].value
    if value is None:
        return value
    return str(value).strip()


def _parse_list(value):
    if value is None:
        return value
    value = value.split(";")
    value = map(str.strip, value)
    value = [obj for obj in value if obj]
    return value


def _check_int(value, output):
    try:
        value = int(value)
    except ValueError:
        print(f"No valid integer given in {output}.")
        sys.exit(1)

    return value


@converter.command(version="0.1.0")
@apy_command
@argument(
    name="file",
    description="The Excel file to read from.",
    param_type=Path(path_type="file", exists=True),
    required=True,
)
@option(
    "force",
    char="f",
    description="Force deleting and overwriting existing information",
    is_flag=True,
)
@option(
    "keep-formulas",
    char="k",
    description="Flag to indicate wether to print the formula of a given cell, even if"
    " a computed value is available.",
    is_flag=True,
)
@option(
    "start-column", char="S", description="Start column", param_type=String, default="E"
)
@option(
    "end-column", char="E", description="End column", param_type=String, default=None
)
@option(
    "permission-new",
    char="p",
    description="Permission of new user",
    default="member",
    param_type=Choice(["member", "editor", "admin"]),
)
@option(
    "base-path",
    char="b",
    description="Prefix path to be added in front of the files or paths"
    " specified in the Excel sheet",
    param_type=String,
    default=None,
)
def excel_to_kadi(
    manager,
    file,
    force,
    keep_formulas,
    start_column,
    end_column,
    base_path,
    permission_new,
):
    """Imports an Excel sheet, reads metadata and transfers them into Kadi."""

    wb = load_workbook(filename=file, read_only=True, data_only=not keep_formulas)
    ws = wb.active

    if not end_column:
        end_column = start_column

    for i in range(ord(start_column), ord(end_column) + 1):

        identifier = _read_value(ws, chr(i), 1)
        title = _read_value(ws, chr(i), 2)
        description = _read_value(ws, chr(i), 3)
        type = _read_value(ws, chr(i), 4)
        tags = _parse_list(_read_value(ws, chr(i), 5))
        add_collections = _parse_list(_read_value(ws, chr(i), 6))
        add_groups = _parse_list(_read_value(ws, chr(i), 7))
        add_user = _parse_list(_read_value(ws, chr(i), 8))
        links = _parse_list(_read_value(ws, chr(i), 10))
        title_links = _parse_list(_read_value(ws, chr(i), 11))
        files = _parse_list(_read_value(ws, chr(i), 13))
        metadatum = _read_value(ws, chr(i), 17)

        if base_path:
            if base_path[-1] != os.sep:
                base_path = base_path + os.sep

        record = manager.cli_record(identifier=identifier, title=title, create=True)

        if description:
            record.set_attribute(description=description)

        if type:
            record.set_attribute(type=type)

        if tags:
            for tag in tags:
                record.add_tag(tag)

        if add_collections:
            for collection_id in add_collections:
                _check_int(collection_id, f"{chr(i)}{6}")
                collection = manager.collection(id=collection_id)
                record.add_collection_link(collection=collection)

        if add_groups:
            for group_id in add_groups:
                _check_int(group_id, f"{chr(i)}{7}")
                group = manager.group(id=group_id)
                record.add_group_role(group=group, permission_new=permission_new)

        if add_user:
            for user_id in add_user:
                _check_int(user, f"{chr(i)}{8}")
                user = manager.user(id=user_id)
                record.add_user(user=user, permission_new=permission_new)

        if files:
            for obj in files:
                if base_path:
                    obj = base_path + obj
                record.upload_file(file_name=obj, force=force, pattern="*")

        if links:
            if len(links) != len(title_links):
                print(
                    f"Found {len(links)} entries for links but {len(title_links)}"
                    " titles. Please use the same number of entries."
                )
                sys.exit(1)
            for link_iter, title_iter in zip(links, title_links):
                _check_int(link_iter, f"{chr(i)}{10}")
                record_to = manager.record(id=link_iter)
                record.link_record(record_to=record_to, name=title_iter)

        if metadatum:
            metadata = []
            x = 17
            while True:
                metadatum_key = _read_value(ws, "A", x)
                if not metadatum_key:
                    break
                metadatum_value = _read_value(ws, chr(i), x)
                metadatum_type = _read_value(ws, "C", x)
                metadatum_unit = _read_value(ws, "D", x)

                metadata.append(
                    validate_metadatum(
                        metadatum=metadatum_key,
                        value=metadatum_value,
                        type=metadatum_type,
                        unit=metadatum_unit,
                    )
                )
                x = x + 1

            record.add_metadata(metadata=metadata, force=force)
