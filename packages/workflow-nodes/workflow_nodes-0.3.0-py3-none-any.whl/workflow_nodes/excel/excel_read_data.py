# Copyright 2020 Karlsruhe Institute of Technology
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import csv
import sys

from openpyxl import load_workbook
from xmlhelpy import argument
from xmlhelpy import Integer
from xmlhelpy import option
from xmlhelpy import String

from .main import excel


@excel.command(version="0.1.0")
@argument(
    name="file",
    description="The Excel file to read from.",
    param_type=String,
    required=True,
)
@option(
    name="column",
    char="c",
    description="The name of the (start) column.",
    param_type=String,
    required=True,
)
@option(
    name="row",
    char="r",
    description="The number of the (start) row.",
    param_type=Integer,
    required=True,
)
@option(
    name="column-end",
    char="C",
    description="The name of the end column.",
    param_type=String,
)
@option(
    name="row-end",
    char="R",
    description="The number of the end row.",
    param_type=Integer,
)
@option(
    "output",
    char="o",
    description="Path to a file to write the output to. If not set, the output will be"
    " written to stdout.",
    param_type=String,
)
@option(
    "keep-formulas",
    char="k",
    description="Flag to indicate wether to print the formula of a given cell, even if"
    " a computed value is available.",
    is_flag=True,
)
def excel_read_data(file, column, row, column_end, row_end, output, keep_formulas):
    """Read data from an Excel file."""

    wb = load_workbook(filename=file, read_only=True, data_only=not keep_formulas)
    ws = wb.active

    if not output:
        output_file = sys.stdout
    else:
        output_file = open(output, "w")

    try:
        if column_end is None and row_end is None:
            value = ws[f"{column}{row}"].value
            if value:
                output_file.write(f"{value}\n")
        else:
            if column_end is None:
                column_end = column

            if row_end is None:
                row_end = row

            csv_writer = csv.writer(
                output_file, delimiter=";", quoting=csv.QUOTE_MINIMAL
            )
            cell_range = ws[f"{column}{row}:{column_end}{row_end}"]

            for cell_row in cell_range:
                csv_writer.writerow([cell.value for cell in cell_row])
    except:
        print(
            "Could not get data from [{}{}].".format(
                f"{column}{row}",
                f":{column_end}{row_end}" if column_end is not None else "",
            )
        )
        sys.exit(1)
