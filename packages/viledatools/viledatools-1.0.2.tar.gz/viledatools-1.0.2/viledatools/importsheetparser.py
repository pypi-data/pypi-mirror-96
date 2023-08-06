# coding=utf-8
"""
(C) FHCS GmbH

Class to parse values from MS Excel sheet given, this class defines and implements MS Excel importing rules
"""

from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from diagonaltablewalk import DiagonalTableWalk
from importcellparser import ImportCellParser
from viledaexceptions import ImportCellParserException, ImportSheetParserException


class ImportSheetParser:
    """
    Class to parse values from MS Excel sheet given, this class defines and implements MS Excel importing rules
    """
    def __init__(self, ws: openpyxl.worksheet.worksheet.Worksheet, wsname: str):
        """
        Gets the imported worksheet object and provides methods for sheet parsing

        :param ws: openpyxl worksheet object to parse
        :param wsname: name of the sheet
        """
        # Local reference to the worksheet
        self._ws = ws
        # Parsed result list of dicts
        self._parsed: List[Dict[str, Any]] = list()
        # Parse log
        self._parse_log = str()
        # Parse OK flag:
        self._parse_ok = True
        # Parsed rows statistics
        self._rows_stats = {"TOTAL": 0, "OK": 0}
        # Row and column of main key tag IQXLSXTABLETAG
        self._mainkey: Tuple[int, int] = (0, 0)
        # Look for the main key
        if not self._find_mainkey():
            raise ImportSheetParserException(f"The main key IQXLSXTABLETAG not found in the excel sheet '{wsname}'")
        # Columns corresponding to field keys in imported MS Excel sheet
        self._columns: Dict[str, int] = dict()
        # Look for column field keys in the same row with main key tag
        self._find_columns()
        if not self._columns:
            # Not found field keys, raise
            raise ImportSheetParserException((f"Found no field key tags in main key row {self._mainkey[0]}, "
                                              f"nothing to import"))
        # Create parsers for each column
        self._p: Dict[str, ImportCellParser] = dict()
        self._register_parsers()
        # Parse
        self._parse_rows()

    def getparsedrecords(self) -> List[Dict[str, Any]]:
        """
        Returns the parsed result

        :return: List of dicts with parsed records for each line
        """
        return self._parsed

    def ok(self) -> bool:
        """
        Returns parse result

        :return: True if all parsed, False if any errors
        """
        return self._parse_ok

    def getlog(self) -> str:
        """
        Returns parse log

        :retrun: parse log
        """
        return self._parse_log

    def getpasrsestats(self) -> Dict[str, int]:
        """
        Returns parse statistics

        :retrun: parse statistics
        """
        return self._rows_stats

    def _find_mainkey(self) -> bool:
        """
        Looks for main key tag IQXLSXTABLETAG in the sheet
        """
        # Diagonal iterator to look for main key tag
        _twalk = DiagonalTableWalk(self._ws.max_row - self._ws.min_row + 1,
                                   self._ws.max_column - self._ws.min_column + 1)
        for _i in _twalk:
            if str(self._ws.cell(self._ws.min_row - 1 + _i[0],
                                 self._ws.min_column - 1 + _i[1]).value).strip() == "IQXLSXTABLETAG":
                # Redefine the self._mainkey with tuple (row, column) of the main key tag
                self._mainkey = (self._ws.min_row - 1 + _i[0], self._ws.min_column - 1 + _i[1])
                self._parse_log += (f"      OK: found main key "
                                    f"at {get_column_letter(self._mainkey[1])}{self._mainkey[0]}\n")
                return True
        return False

    def _find_columns(self):
        """
        Looks for columns containing field keys known to the parser
        """
        _fieldkeys_list = ImportCellParser().getfieldtags()
        for _col in range(self._ws.min_column, self._ws.max_column + 1):
            _col_value = str(self._ws.cell(self._mainkey[0], _col).value).strip()
            if _col_value in _fieldkeys_list:
                if _col_value not in self._columns.keys():
                    self._columns[_col_value] = _col
                    self._parse_log += (f"      OK: found known field key {_col_value} at "
                                        f"{get_column_letter(_col)}{self._mainkey[0]}\n")
                else:
                    self._parse_log += (f"      WARNING: ignored duplicated field key tag {_col_value} "
                                        f"in the cell {get_column_letter(_col)}{self._mainkey[0]}\n")

    def _register_parsers(self):
        """
        Creates parsers for each field key and for each column
        """
        for _field_key, _field_col in self._columns.items():
            try:
                self._p[_field_key] = ImportCellParser(_field_key)
                self._parse_log += f"      OK: registered parser for {_field_key}\n"
            except Exception as _e:
                self._parse_log += f"      ERROR: can not register parser for {_field_key}: {repr(_e)}\n"
                self._parse_ok = False

    def _parse_rows(self):
        """
        Parses all rows containing row tag IQXLSXROWTAG in the main key tag column
        """
        for _row in range(self._mainkey[0] + 1, self._ws.max_row + 1):
            # Get only rows with row key tag
            if str(self._ws.cell(_row, self._mainkey[1]).value) == "IQXLSXROWTAG":
                self._rows_stats["TOTAL"] += 1
                _parsed_row = dict()
                _parsed_ok = True
                for _field_key, _field_col in self._columns.items():
                    _cell_to_parse = self._ws.cell(_row, _field_col)
                    # Provide the openpyxl cell object to corresponding parser
                    try:
                        _parsed_row[_field_key] = self._p[_field_key].parse(_cell_to_parse)
                    except ImportCellParserException as _e:
                        self._parse_log += (f"      ERROR: row {_row} can't parse {_field_key} at "
                                            f"{get_column_letter(_field_col)}{_row}: {repr(_e)}\n")
                        _parsed_ok = False
                if _parsed_ok:
                    self._parsed.append(_parsed_row)
                    self._rows_stats["OK"] += 1
                else:
                    self._parse_log += f"      ERROR: row {_row} was not parsed\n"
